#!/usr/bin/env python3
"""Generate the Cruiser savage gearing page from the source workbook.

The workbook is not a regular table.  This script intentionally models the
source layout:

- Most 7.2/Cruiser rows live under a section title row, not under a repeated
  boss name in every row.
- Some jobs mark the tier with B-column text "至天之座中重量级".
- Other jobs only say "以下为7.2零式配装", then leave B blank on data rows.
- 730 rows must not be treated as Light-heavy by item level because 730 is
  also used by Cloud of Darkness (Chaotic) / 24-man BIS sections.

Run from the repository root:
    python generate_cruiser_gearing.py --write
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any

import openpyxl
from opencc import OpenCC

from convert_urls import convert_url


ROOT = Path(__file__).resolve().parent
XLSX_PATTERN = "FF14*.xlsx"
OUTPUT_MD_BY_TIER = {
    "cruiser": ROOT / "site" / "docs" / "savage" / "cruiser" / "gearing.md",
    "heavy": ROOT / "site" / "docs" / "savage" / "heavy" / "gearing.md",
}
EXPECTED_ROWS = {"cruiser": 53, "heavy": 52}

URL_RE = re.compile(r"https?://[^\s\"]+")
GEAR_LINK_RE = re.compile(r"https://cycleapple\.github\.io/xiv-tc-gear/\?s=[A-Za-z0-9]+")
OPENCC = OpenCC("s2tw")

# Source workbook section labels.
S_HEAVY = "\u81f3\u5929\u4e4b\u5ea7\u91cd\u91cf\u7ea7"
S_CRUISER = "\u81f3\u5929\u4e4b\u5ea7\u4e2d\u91cd\u91cf\u7ea7"
S_LIGHT = "\u81f3\u5929\u4e4b\u5ea7\u8f7b\u91cd\u91cf\u7ea7"
SECTION_BY_LABEL = {S_HEAVY: "heavy", S_CRUISER: "cruiser", S_LIGHT: "light"}

# Headings that change the current tier state.
H_74_RAID = "\u4ee5\u4e0b\u4e3a7.4\u96f6\u5f0f\u914d\u88c5"
H_72_RAID = "\u4ee5\u4e0b\u4e3a7.2\u96f6\u5f0f\u914d\u88c5"

# Non-savage sections in the same sheets.
NON_SAVAGE_MARKERS = (
    "\u706d24",
    "\u706d\u7ea7\u6697\u9ed1\u4e4b\u4e91",
    "\u7edd\u672c",
    "\u963f\u7f57\u5c9b",
    "\u7edd\u6b27",
    "\u9f99\u8bd7",
    "\u7edd\u4e9a",
    "\u795e\u5175",
    "\u5df4\u54c8",
)

CRUISER_ILVLS = {740, 741, 742, 744, 760}

SHEET_DISPLAY = {
    "\u5766\u514b": "\u5766\u514b",
    "\u6cbb\u7597": "\u6cbb\u7642",
    "\u9f99\u9a91": "\u9f8d\u9a0e\u58eb",
    "\u9570\u5200": "\u942e\u5200\u5e2b",
    "\u6b66\u50e7": "\u6b66\u50e7",
    "\u6b66\u58eb": "\u6b66\u58eb",
    "\u8770\u86c7": "\u6bd2\u86c7\u528d\u58eb",
    "\u5fcd\u8005": "\u5fcd\u8005",
    "\u673a\u5de5": "\u6a5f\u5de5\u58eb",
    "\u8bd7\u4eba": "\u541f\u904a\u8a69\u4eba",
    "\u821e\u8005": "\u821e\u8005",
    "\u753b\u5bb6": "\u7e6a\u9748\u6cd5\u5e2b",
    "\u8d64\u9b54": "\u8d64\u9b54\u6cd5\u5e2b",
    "\u53ec\u5524": "\u53ec\u559a\u5e2b",
    "\u9ed1\u9b54": "\u9ed1\u9b54\u6cd5\u5e2b",
}

ROLE_GROUPS = OrderedDict(
    [
        ("\u5766\u514b", ["\u5766\u514b"]),
        ("\u6cbb\u7642", ["\u6cbb\u7597"]),
        (
            "\u8fd1\u6230 DPS",
            ["\u9f99\u9a91", "\u9570\u5200", "\u6b66\u50e7", "\u6b66\u58eb", "\u8770\u86c7", "\u5fcd\u8005"],
        ),
        ("\u9060\u654f DPS", ["\u673a\u5de5", "\u8bd7\u4eba", "\u821e\u8005"]),
        ("\u6cd5\u7cfb DPS", ["\u753b\u5bb6", "\u8d64\u9b54", "\u53ec\u5524", "\u9ed1\u9b54"]),
    ]
)

# Conservative source-term conversion for this gearing sheet.
TERM_REPLACEMENTS = (
    ("\u9490\u9570\u5ba2", "\u942e\u5200\u5e2b"),
    ("\u9490\u9570", "\u942e\u5200\u5e2b"),
    ("\u9570\u5200", "\u942e\u5200\u5e2b"),
    ("\u9f99\u9a91\u58eb", "\u9f8d\u9a0e\u58eb"),
    ("\u9f99\u9a91", "\u9f8d\u9a0e\u58eb"),
    ("\u8770\u86c7\u5251\u58eb", "\u6bd2\u86c7\u528d\u58eb"),
    ("\u8770\u86c7", "\u6bd2\u86c7"),
    ("\u673a\u5de5\u58eb", "\u6a5f\u5de5\u58eb"),
    ("\u673a\u5de5", "\u6a5f\u5de5\u58eb"),
    ("\u8bd7\u4eba", "\u541f\u904a\u8a69\u4eba"),
    ("\u753b\u5bb6", "\u7e6a\u9748\u6cd5\u5e2b"),
    ("\u8d64\u9b54", "\u8d64\u9b54\u6cd5\u5e2b"),
    ("\u53ec\u5524", "\u53ec\u559a\u5e2b"),
    ("\u9ed1\u9b54", "\u9ed1\u9b54\u6cd5\u5e2b"),
    ("\u6cbb\u7597", "\u6cbb\u7642"),
    ("\u5b66\u8005", "\u5b78\u8005"),
    ("\u8d24\u8005", "\u8ce2\u8005"),
    ("\u767d\u8d24", "\u767d\u9b54\uff0f\u8ce2\u8005"),
    ("4H", "4\u6cbb\u7642"),
    ("\u9664\u9a91\u58eb", "\u9664\u9a0e\u58eb"),
    ("\u9a91\u58eb", "\u9a0e\u58eb"),
    ("\u7edd\u67aa", "\u7d55\u69cd"),
    ("\u5f00\u8352", "\u958b\u8352"),
    ("\u6bd5\u4e1a", "\u7562\u696d"),
    ("\u63a8\u8350", "\u63a8\u85a6"),
    ("\u7c7b\u578b", "\u985e\u578b"),
    ("\u804c\u4e1a", "\u8077\u696d"),
    ("\u88c5\u7b49", "\u88dd\u7b49"),
    ("\u70b9\u6570", "\u9ede\u6578"),
    ("\u5151\u6362", "\u514c\u63db"),
    ("\u5151\u63db", "\u514c\u63db"),
    ("\u5151", "\u514c"),
    ("\u6362", "\u63db"),
    ("\u5fc5\u9009", "\u5fc5\u9078"),
    ("\u53ef\u9009", "\u53ef\u9078"),
    ("\u66ff\u6362", "\u66ff\u63db"),
    ("\u51c6\u5907\u5468", "\u6e96\u5099\u9031"),
    ("\u51c6\u5907", "\u6e96\u5099"),
    ("\u96f6\u5f0f\u5468", "\u96f6\u5f0f\u9031"),
    ("\u6bd5\u4e1a\u5468\u6570", "\u7562\u696d\u9031\u6578"),
    ("\u5468", "\u9031"),
    ("\u9879\u94fe", "\u9805\u934a"),
    ("\u9879\u73af", "\u9805\u74b0"),
    ("\u8033\u5760", "\u8033\u589c"),
    ("\u8033\u9970", "\u8033\u98fe"),
    ("\u8033\u5939", "\u8033\u593e"),
    ("\u624b\u956f", "\u624b\u9432"),
    ("\u88e4\u5b50", "\u8932\u5b50"),
    ("\u9996\u9970", "\u9996\u98fe"),
    ("\u9970\u54c1", "\u98fe\u54c1"),
    ("\u66b4\u51fb", "\u66b4\u64ca"),
    ("\u76f4\u51fb", "\u76f4\u64ca"),
    ("\u548f\u5531", "\u8a60\u5531"),
    ("\u7981\u65ad", "\u7981\u65b7"),
    ("\u9576\u5d4c", "\u9472\u5d4c"),
    ("\u5c5e\u6027", "\u5c6c\u6027"),
    ("\u517c\u5bb9\u6027", "\u76f8\u5bb9\u6027"),
    ("\u517c\u5bb9", "\u76f8\u5bb9"),
    ("\u5f53\u524d", "\u76ee\u524d"),
    ("\u4e24\u4e2a", "\u5169\u500b"),
    ("\u4e24\u8005", "\u5169\u8005"),
    ("\u65e0\u635f", "\u7121\u640d"),
    ("\u5c42\u65f6", "\u5c64\u6642"),
    ("\u8fdc\u654f", "\u9060\u654f"),
    ("\u9f99\u9570", "\u9f8d\u942e"),
    ("\u8be5", "\u8a72"),
    ("\u4e0e", "\u8207"),
    ("\u65e0", "\u7121"),
    ("\u5e76", "\u4e26"),
    ("\u540e", "\u5f8c"),
    ("\u4e3a", "\u70ba"),
    ("\u4e2a", "\u500b"),
    ("\u9897", "\u9846"),
    ("\u4e13\u7528", "\u5c08\u7528"),
    ("\u5f3a\u884c", "\u5f37\u884c"),
    ("\u5e38\u89c4", "\u5e38\u898f"),
    ("\u5c06", "\u5c07"),
    ("\u4f18\u5316", "\u512a\u5316"),
    ("\u4f18\u52bf", "\u512a\u52e2"),
    ("\u4e07", "\u842c"),
    ("\u9f99", "\u9f8d"),
    ("\u9570", "\u942e"),
    ("\u4e13", "\u5c08"),
    ("\u5f3a", "\u5f37"),
    ("\u548f", "\u8a60"),
    ("\u88c5", "\u88dd"),
    ("\u6ca1", "\u6c92"),
    ("\u753b", "\u7e6a"),
    ("\u70b9", "\u9ede"),
    ("\u540c\u65f6", "\u540c\u6642"),
    ("\u9ad8\u548f", "\u9ad8\u8a60"),
    ("\u7406\u8bba", "\u7406\u8ad6"),
    ("\u8fc7\u5ea6", "\u904e\u6e21"),
    ("\u8fc7\u6e21", "\u904e\u6e21"),
    ("\u9002\u5e94", "\u9069\u61c9"),
    ("\u6570\u503c", "\u6578\u503c"),
    ("\u8def\u7ebf", "\u8def\u7dda"),
    ("\u7cbe\u51c6", "\u7cbe\u6e96"),
    ("\u548f\u5492", "\u8a60\u5492"),
    ("\u77f3\u5934", "\u77f3\u982d"),
)

DISALLOWED_PROSE_TERMS = (
    "\u5151",
    "\u5c42",
    "\u4e13",
    "\u5f3a",
    "\u89c4",
    "\u5c06",
    "\u6362",
    "\u4f18",
    "\u4e07",
    "\u8fdc",
    "\u9970",
    "\u51c6\u5907",
    "\u9897",
    "\u65e0\u635f",
    "\u5e76",
    "\u8fc7\u6e21",
    "\u548f",
    "\u9f99",
    "\u9570",
    "\u88c5",
    "\u6ca1",
    "\u753b",
    "\u70b9",
    "\u8fc7\u5ea6",
    "\u94c1",
    "\u9430",
)


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def markdown_text(value: Any) -> str:
    text = as_text(value).replace("\r\n", "\n").replace("\r", "\n")
    for source, target in TERM_REPLACEMENTS:
        text = text.replace(source, target)
    text = OPENCC.convert(text)
    text = text.replace("|", "\uff0f")
    lines = [part.strip() for part in text.split("\n") if part.strip()]
    return "<br />".join(lines)


def urls_from(value: Any) -> list[str]:
    if value is None:
        return []
    return [url.rstrip("),") for url in URL_RE.findall(str(value))]


def first_source_url(ws: Any, row: int) -> str:
    # Prefer the raw source columns.  I/J are formulas that derive display links
    # from L/M and can include the alternate domain.
    urls: list[str] = []
    for col in (12, 13):
        urls.extend(urls_from(ws.cell(row, col).value))
    if urls:
        return urls[0]

    fallback: list[str] = []
    for col in range(9, min(ws.max_column, 15) + 1):
        fallback.extend(urls_from(ws.cell(row, col).value))
    return fallback[0] if fallback else ""


def item_level(value: Any) -> int | None:
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"\d+", str(value or ""))
    return int(match.group(0)) if match else None


def extract_entries(workbook_path: Path, target_tier: str) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)
    entries: list[dict[str, Any]] = []

    for ws in workbook.worksheets[2:17]:
        current_tier: str | None = None
        explicit_tier: str | None = None

        for row in range(1, ws.max_row + 1):
            row_text = " ".join(as_text(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 15) + 1))
            b_value = ws.cell(row, 2).value
            b_text = b_value.strip() if isinstance(b_value, str) else ""

            if b_text in SECTION_BY_LABEL:
                explicit_tier = SECTION_BY_LABEL[b_text]
                current_tier = explicit_tier
            elif H_74_RAID in row_text:
                current_tier = "heavy"
                explicit_tier = None
            elif H_72_RAID in row_text:
                current_tier = "cruiser"
                explicit_tier = None
            elif any(marker in row_text for marker in NON_SAVAGE_MARKERS):
                if "\u914d\u88c5" in row_text or b_text:
                    current_tier = None
                    explicit_tier = None

            source_url = first_source_url(ws, row)
            type_value = ws.cell(row, 4).value
            gcd_value = ws.cell(row, 5).value
            job_value = ws.cell(row, 6).value
            if not (source_url and (type_value or gcd_value or job_value)):
                continue

            ilvl = item_level(ws.cell(row, 3).value)
            tier = explicit_tier or current_tier
            if tier is None and ilvl in CRUISER_ILVLS:
                tier = "cruiser"
            if tier == "heavy" and ilvl in CRUISER_ILVLS:
                tier = "cruiser"

            if tier != target_tier:
                continue

            entries.append(
                {
                    "sheet": ws.title,
                    "row": row,
                    "ilvl": as_text(ws.cell(row, 3).value),
                    "type": markdown_text(type_value),
                    "gcd": markdown_text(gcd_value),
                    "job": markdown_text(job_value),
                    "note": markdown_text(ws.cell(row, 11).value),
                    "source_url": source_url,
                }
            )

    deduped: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for entry in entries:
        if entry["source_url"] in seen_urls:
            continue
        seen_urls.add(entry["source_url"])
        deduped.append(entry)
    return deduped


def extract_cruiser_entries(workbook_path: Path) -> list[dict[str, Any]]:
    return extract_entries(workbook_path, "cruiser")


def convert_entries(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    converted: list[dict[str, Any]] = []
    for index, entry in enumerate(entries, 1):
        print(
            f"{index:02d}/{len(entries)} {entry['sheet']}:{entry['row']} "
            f"{entry['type']} {entry['gcd']} {entry['job']}",
            flush=True,
        )
        result = dict(entry)
        result["link"] = convert_url(entry["source_url"])
        converted.append(result)
    return converted


TIER_META = {
    "cruiser": {
        "label": "\u4e2d\u91cd\u91cf\u7d1a",
        "range": "M5S\uff5eM8S",
        "source": "7.2 \u96f6\u5f0f\u914d\u88dd\uff0f\u81f3\u5929\u4e4b\u5ea7\u4e2d\u91cd\u91cf\u7d1a",
        "heading": "\u81f3\u5929\u4e4b\u5ea7\u4e2d\u91cd\u91cf\u7d1a\uff08M5S\uff5eM8S\uff09",
        "show_ilvl": True,
        "show_note": True,
    },
    "heavy": {
        "label": "\u91cd\u91cf\u7d1a",
        "range": "M9S\uff5eM12S",
        "source": "7.4 \u96f6\u5f0f\u914d\u88dd\uff0f\u81f3\u5929\u4e4b\u5ea7\u91cd\u91cf\u7d1a",
        "heading": "\u81f3\u5929\u4e4b\u5ea7\u91cd\u91cf\u7d1a\uff08M9S\uff5eM12S\uff09",
        "show_ilvl": False,
        "show_note": True,
    },
}


def render_markdown(entries: list[dict[str, Any]], tier: str) -> str:
    by_sheet: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for entry in entries:
        by_sheet.setdefault(entry["sheet"], []).append(entry)

    meta = TIER_META[tier]
    lines: list[str] = [
        "---",
        "sidebar_position: 1",
        "title: \u914d\u88dd\u63a8\u85a6",
        f"description: FF14 \u963f\u5361\u72c4\u4e9e\u96f6\u5f0f{meta['label']}\uff08{meta['range']}\uff09\u5168\u8077\u696d\u914d\u88dd\u63a8\u85a6 \u2014 \u5766\u514b\u3001\u6cbb\u7642\u3001DPS \u958b\u8352\u8207\u7562\u696d BIS",
        "---",
        "",
        "# \u914d\u88dd\u63a8\u85a6",
        "",
        ":::info \u914d\u88dd\u4f86\u6e90",
        f"\u914d\u88dd\u65b9\u6848\u53c3\u8003\u81ea\u55b5\u53d4\u5b64\u98a8\u884c\u6574\u7406\u7684[FF14\u5168\u6230\u8077\u914d\u88dd\u8868](https://www.kdocs.cn/l/ceEcTzlFQBUy)\u3002\u672c\u9801\u7531\u8a66\u7b97\u8868\u4e2d\u300c{meta['source']}\u300d\u5340\u6bb5\u7522\u751f\uff0c\u9ede\u64ca\u300c\u67e5\u770b\u914d\u88dd\u300d\u53ef\u5728\u914d\u88dd\u5de5\u5177\u4e2d\u67e5\u770b\u5b8c\u6574\u88dd\u5099\u8207\u9b54\u6676\u77f3\u8a73\u60c5\u3002",
        ":::",
        "",
        "| \u8077\u80fd | \u8077\u696d |",
        "|------|------|",
        "| \u5766\u514b | [\u9632\u8b77\u8077\u80fd\u901a\u7528](#\u5766\u514b) |",
        "| \u6cbb\u7642 | [\u6cbb\u7642\u8077\u80fd\u901a\u7528](#\u6cbb\u7642) |",
        "| \u8fd1\u6230 | [\u9f8d\u9a0e\u58eb](#\u9f8d\u9a0e\u58eb) \u00b7 [\u942e\u5200\u5e2b](#\u942e\u5200\u5e2b) \u00b7 [\u6b66\u50e7](#\u6b66\u50e7) \u00b7 [\u6b66\u58eb](#\u6b66\u58eb) \u00b7 [\u6bd2\u86c7\u528d\u58eb](#\u6bd2\u86c7\u528d\u58eb) \u00b7 [\u5fcd\u8005](#\u5fcd\u8005) |",
        "| \u9060\u654f | [\u6a5f\u5de5\u58eb](#\u6a5f\u5de5\u58eb) \u00b7 [\u541f\u904a\u8a69\u4eba](#\u541f\u904a\u8a69\u4eba) \u00b7 [\u821e\u8005](#\u821e\u8005) |",
        "| \u6cd5\u7cfb | [\u7e6a\u9748\u6cd5\u5e2b](#\u7e6a\u9748\u6cd5\u5e2b) \u00b7 [\u8d64\u9b54\u6cd5\u5e2b](#\u8d64\u9b54\u6cd5\u5e2b) \u00b7 [\u53ec\u559a\u5e2b](#\u53ec\u559a\u5e2b) \u00b7 [\u9ed1\u9b54\u6cd5\u5e2b](#\u9ed1\u9b54\u6cd5\u5e2b) |",
        "",
        "---",
        "",
        f"## {meta['heading']}",
        "",
    ]

    for role, sheets in ROLE_GROUPS.items():
        if not any(sheet in by_sheet for sheet in sheets):
            continue
        lines.extend([f"### {role}", ""])

        for sheet in sheets:
            rows = by_sheet.get(sheet)
            if not rows:
                continue

            if role in {"\u8fd1\u6230 DPS", "\u9060\u654f DPS", "\u6cd5\u7cfb DPS"}:
                header = (
                    "| \u88dd\u7b49 | \u985e\u578b | GCD | \u914d\u88dd | \u5099\u8a3b |"
                    if meta["show_ilvl"]
                    else "| \u985e\u578b | GCD | \u914d\u88dd | \u5099\u8a3b |"
                )
                rule = (
                    "|------|------|-----|------|------|"
                    if meta["show_ilvl"]
                    else "|------|-----|------|------|"
                )
                lines.extend(
                    [
                        f"#### {SHEET_DISPLAY[sheet]}",
                        "",
                        header,
                        rule,
                    ]
                )
                for row in rows:
                    cells = (
                        [row["ilvl"], row["type"], row["gcd"]]
                        if meta["show_ilvl"]
                        else [row["type"], row["gcd"]]
                    )
                    cells.extend([f"[\u67e5\u770b\u914d\u88dd]({row['link']})", row["note"]])
                    lines.append("| " + " | ".join(cells) + " |")
                lines.append("")
            else:
                header = (
                    "| \u8077\u696d | \u88dd\u7b49 | \u985e\u578b | GCD | \u914d\u88dd | \u5099\u8a3b |"
                    if meta["show_ilvl"]
                    else "| \u8077\u696d | \u985e\u578b | GCD | \u914d\u88dd | \u5099\u8a3b |"
                )
                rule = (
                    "|------|------|------|-----|------|------|"
                    if meta["show_ilvl"]
                    else "|------|------|-----|------|------|"
                )
                lines.extend(
                    [
                        header,
                        rule,
                    ]
                )
                for row in rows:
                    cells = (
                        [row["job"], row["ilvl"], row["type"], row["gcd"]]
                        if meta["show_ilvl"]
                        else [row["job"], row["type"], row["gcd"]]
                    )
                    cells.extend([f"[\u67e5\u770b\u914d\u88dd]({row['link']})", row["note"]])
                    lines.append("| " + " | ".join(cells) + " |")
                lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def validate(entries: list[dict[str, Any]], markdown: str, tier: str) -> None:
    errors: list[str] = []

    expected_rows = EXPECTED_ROWS[tier]
    if len(entries) != expected_rows:
        errors.append(f"expected {expected_rows} {tier} rows, got {len(entries)}")

    source_urls = [entry["source_url"] for entry in entries]
    if len(source_urls) != len(set(source_urls)):
        errors.append("duplicate source URLs detected")

    generated_links = [entry["link"] for entry in entries]
    markdown_links = GEAR_LINK_RE.findall(markdown)
    if generated_links != markdown_links:
        errors.append(
            f"markdown links do not match converted source order: entries={len(generated_links)}, md={len(markdown_links)}"
        )

    if "\ufffd" in markdown:
        errors.append("replacement character U+FFFD found in markdown")

    # Ignore the query separator in URLs; prose should not contain double
    # question marks after generation.
    prose_without_urls = GEAR_LINK_RE.sub("", markdown)
    if "??" in prose_without_urls:
        errors.append("double question marks found outside generated URLs")

    residual_terms = [term for term in DISALLOWED_PROSE_TERMS if term in prose_without_urls]
    if residual_terms:
        errors.append(
            "unconverted source terms found in markdown prose: "
            + ", ".join(repr(term) for term in residual_terms)
        )

    if errors:
        raise SystemExit("Validation failed:\n- " + "\n- ".join(errors))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--tier",
        choices=["cruiser", "heavy", "all"],
        default="cruiser",
        help="which tier to generate",
    )
    parser.add_argument("--write", action="store_true", help="write generated markdown files")
    parser.add_argument("--json", type=Path, help="optional path to write extracted entries as JSON")
    args = parser.parse_args()

    candidates = sorted(ROOT.glob(XLSX_PATTERN))
    if not candidates:
        raise SystemExit(f"No workbook matching {XLSX_PATTERN!r} found in {ROOT}")
    workbook_path = candidates[0]

    tiers = ["cruiser", "heavy"] if args.tier == "all" else [args.tier]
    json_payload: dict[str, list[dict[str, Any]]] = {}
    rendered: dict[str, str] = {}

    for tier in tiers:
        entries = extract_entries(workbook_path, tier)
        print(f"Extracted {tier} entries: {len(entries)} from {workbook_path.name}", flush=True)
        converted = convert_entries(entries)
        markdown = render_markdown(converted, tier)
        validate(converted, markdown, tier)
        json_payload[tier] = converted
        rendered[tier] = markdown

    if args.json:
        payload: Any = json_payload[tiers[0]] if len(tiers) == 1 else json_payload
        args.json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"Wrote {args.json}", flush=True)

    if args.write:
        for tier, markdown in rendered.items():
            output_md = OUTPUT_MD_BY_TIER[tier]
            output_md.write_text(markdown, encoding="utf-8")
            print(f"Wrote {output_md}", flush=True)
    else:
        print("\n".join(rendered.values()))


if __name__ == "__main__":
    main()
